import json
import os
from re import search
import openpyxl
import pycountry
import requests
from textblob import TextBlob
from openpyxl.styles import Font, Color, colors, fills


# ---------------------------------
# Created By: Juvon Hyatt
# ----------------------------------

# How to use this program:
# 1) fill AutoBatchPrint.py sheet[template] with data where and how the template indicates
# 2) fill AutoBatchPrint.py sheet[address] with the address as specified
# 3) Run
# 4) Save finished template under a different name and clear out the FinalShippingTemplate for future use

global vfstreet
global vfstreet2
global vfcity
global vfstate
global vfzipcode

global address1
global address2
global city
global state
global zipcode
global found

class Batch:
    def __init__(self, file_name):
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook["template"]
        self.print_type = sheet.cell(row=2, column=2).value
        self.num_packages = sheet.cell(row=3, column=2).value
        self.Ref1 = sheet.cell(row=4, column=2).value
        self.PackageId = sheet.cell(row=5, column=2).value
        self.residential = sheet.cell(row=6, column=2).value
        self.Carrier = sheet.cell(row=7, column=2).value
        self.Method = sheet.cell(row=8, column=2).value
        self.Shipping_Profile = sheet.cell(row=9, column=2).value
        self.Billing = sheet.cell(row=10, column=2).value
        self.AccountNum = sheet.cell(row=11, column=2).value
        self.TP_BCName = sheet.cell(row=12, column=2).value
        self.TP_Address = sheet.cell(row=13, column=2).value
        self.TP_City = sheet.cell(row=14, column=2).value
        self.TP_State = sheet.cell(row=15, column=2).value
        self.TP_Zip = sheet.cell(row=16, column=2).value
        self.Sat_Del = sheet.cell(row=17, column=2).value
        self.Package_type = sheet.cell(row=18, column=2).value
        self.Confirmation = sheet.cell(row=19, column=2).value
        self.SF_Company = sheet.cell(row=20, column=2).value
        self.SF_Address1 = sheet.cell(row=21, column=2).value
        self.SF_Address2 = sheet.cell(row=22, column=2).value
        self.SF_City = sheet.cell(row=23, column=2).value
        self.SF_State = sheet.cell(row=24, column=2).value
        self.SF_Zip = sheet.cell(row=25, column=2).value
        self.SF_Phone = sheet.cell(row=26, column=2).value
        self.Weight = sheet.cell(row=27, column=2).value
        self.Length = sheet.cell(row=28, column=2).value
        self.Width = sheet.cell(row=29, column=2).value
        self.Height = sheet.cell(row=30, column=2).value
        self.Description = sheet.cell(row=34, column=3).value
        self.Price = sheet.cell(row=35, column=3).value
        self.Origin = sheet.cell(row=36, column=3).value
        self.Harmonized_Code = sheet.cell(row=37, column=3).value
        workbook.close()

    def display(self):
        print(f' ------------{self.print_type}------------')
        print(f'Number Of packages: {self.num_packages}')
        print(f'Ref1: {self.Ref1}')
        print(f'Package ID: {self.PackageId}')
        print(f'Residential: {self.residential}')
        print(f'Carrier: {self.Carrier}')
        print(f'Shipping Profile: {self.Shipping_Profile}')
        print(f'Method: {self.Method}')
        print(f'Billing: {self.Billing}')
        print(f'Account Number: {self.AccountNum}')
        print(f'Third Party Name: {self.TP_BCName}')
        print(f'Third Party Address: {self.TP_Address}')
        print(f'Third Party State: {self.TP_State}')
        print(f'Third Party Zip: {self.TP_Zip}')
        print(f'Saturday Delivery: {self.Sat_Del}')
        print(f'Package Type: {self.Package_type}')
        print(f'Confirmation: {self.Confirmation}')
        print(f'Sent From Company: {self.SF_Company}')
        print(f'Sent From Address: {self.SF_Address1}')
        print(f'Sent From Address2: {self.SF_Address2}')
        print(f'Sent From City: {self.SF_City}')
        print(f'Sent From State: {self.SF_State}')
        print(f'Sent From Phone #: {self.SF_Phone}')
        print(f'Weight: {self.Weight}')
        print(f'Length: {self.Length}')
        print(f'Width: {self.Width}')
        print(f'Height: {self.Height}')
        print(f'General Description Of Goods: {self.Description}')
        print(f'Unit Price: {self.Price}')
        print(f'Harmonized Code: {self.Harmonized_Code}')
        print(f'Origin Country: {self.Origin}')
        print('------------------------------')


# -----------------  DATA CORRECTION -----------------------------
# corrects spelling of misspelled words
def auto_correct(text):  # Opening the test file with the intention to read
    textBlb = TextBlob(text)  # Making our first textblob
    textCorrected = textBlb.correct()
    if len(text) <= 2:  # this is trying to catch 2-letter codes, not being autocorrected into words
        textCorrected = text
    return str(textCorrected)  # Correcting the text


# Converts country name to 2-letter country code
def country_search(country):
    try:
        result = pycountry.countries.search_fuzzy(auto_correct(country))
    except Exception:
        return country
    else:
        return result[0].alpha_2


# This fixes zips in the US that have less than 5 places
def fix_zip(zpc):
    return str(zpc).zfill(5)  # adds leading 5 0's

def phone_number_format(num):
    num = num.replace('(', '').replace(')', '').replace('-', '')

    if len(num) > 10:
        return num[1:]
    else:
        return num  

# delete previous data
def purge(sheet):
    pass


# US state to 2-letter code
us_state_to_abbrev = {
    "alabama": "AL",
    "alaska": "AK",
    "arizona": "AZ",
    "arkansas": "AR",
    "california": "CA",
    "colorado": "CO",
    "connecticut": "CT",
    "delaware": "DE",
    "florida": "FL",
    "georgia": "GA",
    "hawaii": "HI",
    "idaho": "ID",
    "illinois": "IL",
    "indiana": "IN",
    "iowa": "IA",
    "kansas": "KS",
    "kentucky": "KY",
    "louisiana": "LA",
    "maine": "ME",
    "maryland": "MD",
    "massachusetts": "MA",
    "michigan": "MI",
    "minnesota": "MN",
    "mississippi": "MS",
    "missouri": "MO",
    "montana": "MT",
    "nebraska": "NE",
    "nevada": "NV",
    "new hampshire": "NH",
    "new jersey": "NJ",
    "new mexico": "NM",
    "new york": "NY",
    "north carolina": "NC",
    "north dakota": "ND",
    "ohio": "OH",
    "oklahoma": "OK",
    "oregon": "OR",
    "pennsylvania": "PA",
    "rhode island": "RI",
    "south carolina": "SC",
    "south dakota": "SD",
    "tennessee": "TN",
    "texas": "TX",
    "utah": "UT",
    "vermont": "VT",
    "virginia": "VA",
    "washington": "WA",
    "west virginia": "WV",
    "wisconsin": "WI",
    "wyoming": "WY",
    "district of columbia": "DC",
    "american samoa": "AS",
    "guam": "GU",
    "northern mariana islands": "MP",
    "puerto rico": "PR",
    "united states minor outlying islands": "UM",
    "u.s. virgin islands": "VI",
}


# ---------------- Address verification SMARTY ----------------------

def jprint(obj):
    # create a formatted string of the Python JSON object
    text = json.dumps(obj, sort_keys=True, indent=4)
    print(text)


def verify_address_US(street, street2, city, state, zipcode):

    street.replace(" ", "+")
    street2.replace(" ", "+")
    street_address = street + ' ' + street2
    city.replace(" ", "+")
    state.replace(" ", "+")
    zipcode.replace(" ", "+")

    codes = {
        '200': 'Everything went okay, and the result has been returned (if any).',
        '301': 'The server is redirecting you to a different endpoint. This can happen when a company switches domain names, or an endpoint name is changed.',
        '400': 'The server thinks you made a bad request. This can happen when you don’t send along the right data, among other things.',
        '401': 'The server thinks you’re not authenticated. Many APIs require login credentials, so this happens when you don’t send the right credentials to access an API.',
        '403': 'The resource you’re trying to access is forbidden: you don’t have the right permissions to see it.',
        '404': 'The resource you tried to access wasn’t found on the server.',
        '503': 'The server is not ready to handle the request.'
    }

    key = ''
    auth_id = ''
    auth_token = ''

    req = f'https://us-street.api.smartystreets.com/street-address?auth-id={auth_id}&auth-token={auth_token}&street=' \
          f'{street_address}&city={city}&state={state}&candidates=10'
    req.replace(" ", '%')
    response = requests.get(req)

    response_code = str(response).split()[1].replace('[', '').replace(']', '').replace('>', '')
    # print(f'{error}- {codes.get(error)}')
    # print(f'Req: {req}')

    # loading json data as text dictionary
    response_data = json.loads(response.text)
    print(response_data[0]['errors'])
    if len(response_data) != 0:

        # jprint(response.json())

        # corrected data parsing
        vstreet = f"{response_data[0]['components']['primary_number']} " \
                  f"{response_data[0]['components']['street_name']}"
        print(vstreet)
        try:
            vstreet_suffix = response_data[0]['components']['street_suffix']
        except KeyError:
            vstreet_suffix = ''

        vstreet = f"{vstreet} {vstreet_suffix}"

        if street2 != '':
            try:
                vstreet2 = f"{response_data[0]['components']['secondary_designator']} " \
                           f"{response_data[0]['components']['secondary_number']}"
            except KeyError:
                vstreet2 = f"{response_data[0]['components']['secondary_designator']}"
        else:
            vstreet2 = ''

        vcity = response_data[0]['components']['city_name']
        vstate = response_data[0]['components']['state_abbreviation']
        vzip = response_data[0]['components']['zipcode']

        if (street.lower() == vstreet.lower()) and (city.lower() == vcity.lower()) \
                and (state.lower() == vstate.lower()) and (zipcode.lower() == vzip.lower()):
            print("--------------\nValid Address\n--------------")
        else:
            print("--------------\nCorrected Address\n--------------")
        print(f"{vstreet} {vstreet2}, {vcity}, {vstate} {vzip}")
        return vstreet, vstreet2, vcity, vstate, vzip

    else:
        pass
        # print("--------------\nNo Response\n--------------")



# ---------------- TEMPLATE PROCESSING -----------------------

def fill_data_ups(batch):
    print("Loading Workbook...")
    workbook = openpyxl.load_workbook("FinalShippingTemplate.xlsx")
    workbook2 = openpyxl.load_workbook("AutoBatchPrintTemplate.xlsx")
    sheet = workbook["UPS"]
    address = workbook2["Address"]
    row = 19
    arow = 2
    count = 1  # where the AI will begin
    print('got', batch)
    print(f'{batch.num_packages} rows expected')

    # process_addresses = input("Verify Addresses?")
    process_addresses = 'no'

    for i in range(0, int(batch.num_packages)):
        batch.Billing = batch.Billing.lower() if batch.Billing is not None else batch.Billing
        batch.residential = batch.residential.lower() if batch.residential is not None else batch.residential
        batch.Description = batch.Description.lower() if batch.Description is not None else batch.Description
        batch.Method = batch.Method.lower() if batch.Method is not None else batch.Method
        ShipToCountry = address.cell(row=arow, column=8).value

        # -------------------------------------VERIFY ADDRESS -------------------------------------------------------
        address1 = address.cell(row=arow, column=2).value  # Address
        address2 = address.cell(row=arow, column=3).value if address.cell(row=arow, column=3).value is not None else ''
        city = address.cell(row=arow, column=5).value  # City
        country = country_search(address.cell(row=arow, column=8).value)  # search, convert the country to 2-letter code
        state = us_state_to_abbrev.get(address.cell(row=arow, column=6).value.strip().lower(),
                                       address.cell(row=arow,
                                                    column=6).value.strip()) if country == 'US' else (
            address.cell(row=arow, column=6).value.strip() if address.cell(row=arow,
                                                                           column=6).value is not None else '')
        zipcode = address.cell(row=arow, column=7).value if country != 'US' else fix_zip(
            address.cell(row=arow, column=7).value)

        found = True
        if process_addresses.lower() == 'yes':
            # runs US address verification only if the address is in the US
            if country == 'US':
                try:
                    vfstreet, vfstreet2, vfcity, vfstate, vfzipcode = verify_address_US(address1, address2, city, state,
                                                                                        zipcode)

                    if vfstreet is not None:
                        address1 = vfstreet
                        address2 = vfstreet2
                        city = vfcity
                        state = vfstate
                        zipcode = vfzipcode
                        print("Address Found")
                    else:
                        found = False
                except:
                    pass

                # color unverified rows
                if not found:
                    sheet[f'A{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'B{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'C{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'D{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'E{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'F{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'G{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'H{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'I{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'J{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'L{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'M{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'N{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'O{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))

                    sheet[f'P{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'Q{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'R{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'S{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'T{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'U{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))

                    sheet[f'V{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'W{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'X{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'Y{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))

                    sheet[f'Z{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AA{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AC{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AD{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AE{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AF{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AG{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AH{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AI{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AJ{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AK{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AL{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AM{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AN{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AO{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AP{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AQ{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AR{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AS{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AT{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AU{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AV{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AW{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AX{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AY{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AZ{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
        # ------------------------------------ WRITING THE DATA--------------------------------------------------------

        if address.cell(row=arow, column=14).value is None:  # custom sort identifier
            sheet[f'A{row}'] = f'{batch.PackageId}-{str(count).zfill(4)}'
            sheet[f'B{row}'] = f'{batch.PackageId}-{str(count).zfill(4)}'
        else:
            sheet[f'A{row}'] = f'({address.cell(row=arow, column=14).value}){batch.PackageId}-{str(count).zfill(4)}'
            sheet[f'B{row}'] = f'({address.cell(row=arow, column=14).value}){batch.PackageId}-{str(count).zfill(4)}'

        sheet[f'C{row}'] = address.cell(row=arow, column=1).value.strip() if address.cell(row=arow, column=1
                                                                                          ).value is not None else ''  # company
        sheet[f'D{row}'] = ''
        sheet[f'E{row}'] = address1
        sheet[f'F{row}'] = address2
        sheet[f'G{row}'] = str(address.cell(row=arow, column=4).value).strip() if address.cell(row=arow, column=4
                                                                                               ).value is not None else ''  # Address3
        sheet[f'H{row}'] = city
        country = country_search(address.cell(row=arow, column=8).value).strip() if country_search(
            address.cell(row=arow, column=8).value) is not None else country_search(
            address.cell(row=arow, column=8).value)  # Search, Convert country to
        # 2-letter code
        sheet[f'I{row}'] = state
        sheet[f'K{row}'] = 'Y' if batch.residential == 'yes' else 'N'
        sheet[f'L{row}'] = batch.Ref1.strip()

        if address.cell(row=arow, column=14).value is not None:  # custom sort identifier
            sheet[f'M{row}'] = f'({address.cell(row=arow, column=14).value}){batch.PackageId}-{str(count).zfill(4)}'.strip()
        else:
            sheet[f'M{row}'] = f'{batch.PackageId}-{str(count).zfill(4)}'.strip()

        sheet[f'J{row}'] = zipcode
        sheet[f'N{row}'] = country  # ship to country
        sheet[f'O{row}'] = batch.Method.strip()
        sheet[f'P{row}'] = 'TP' if batch.Billing.strip() == 'third party' else batch.Billing
        sheet[f'Q{row}'] = batch.Package_type.strip()
        sheet[f'R{row}'] = batch.Weight if float(batch.Weight) >= 1 else '1'
        sheet[f'S{row}'] = batch.SF_Company.strip()
        sheet[f'T{row}'] = 'Shipper'
        sheet[f'U{row}'] = batch.SF_Address1.strip()
        sheet[f'V{row}'] = 'US'
        sheet[f'W{row}'] = batch.SF_Zip
        sheet[f'X{row}'] = batch.SF_City.strip()
        sheet[f'Y{row}'] = (us_state_to_abbrev.get(batch.SF_State.lower().strip())) if (
                    batch.SF_State is not None and len(batch.SF_State) != 2) else batch.SF_State
        sheet[f'Z{row}'] = batch.SF_Phone

        # sheet[f'AA{row}'] =  # id
        # sheet[f'AA{row}'] =  # id
        # sheet[f'AB{row}'] =  # id
        # sheet[f'AC{row}'] =  # id
        sheet[f'AD{row}'] = batch.Sat_Del
        sheet[f'AE{row}'] = phone_number_format(str(address.cell(row=arow, column=10).value)) if address.cell(row=arow, column=10).value is not None else '' # ship to phone
        sheet[f'AF{row}'] = batch.Length
        sheet[f'AG{row}'] = batch.Width
        sheet[f'AH{row}'] = batch.Height
        sheet[f'AI{row}'] = batch.TP_BCName.strip() if batch.TP_BCName is not None else batch.TP_BCName
        sheet[f'AJ{row}'] = batch.TP_Address.strip() if batch.TP_Address is not None else batch.TP_Address
        sheet[f'AK{row}'] = '' if batch.Billing == 'prepaid' else 'United States'
        sheet[f'AL{row}'] = batch.TP_Zip
        sheet[f'AM{row}'] = batch.TP_City.strip() if batch.TP_City is not None else batch.TP_City
        sheet[f'AN{row}'] = batch.TP_State.strip() if batch.TP_State is not None else batch.TP_State
        sheet[f'AO{row}'] = batch.AccountNum
        sheet[f'AP{row}'] = batch.Description.strip() if country != 'US' else ''
        sheet[f'AQ{row}'] = '1'
        sheet[f'AR{row}'] = 'EA'
        sheet[f'AS{row}'] = batch.Price if country != 'US' else ''
        sheet[f'AT{row}'] = 'USD'
        sheet[f'AU{row}'] = 'Shipper' if batch.Billing == 'prepaid' or batch.Billing == 'PREPAID' else 'TP'
        sheet[f'AV{row}'] = 'Shipper' if batch.Billing == 'prepaid' or batch.Billing == 'PREPAID' else 'TP'
        # sheet[f'AW{row}'] =  # id
        # sheet[f'AX{row}'] =  # id
        sheet[f'AY{row}'] = 'United States' if batch.Origin == '' else batch.Origin
        sheet[f'AZ{row}'] = address.cell(row=arow, column=9).value  # email
        sheet[f'BI{row}'] = batch.Harmonized_Code

        row += 1
        arow += 1
        count += 1

    workbook.save("FinalShippingTemplate.xlsx")
    print("Finished")

def fill_data_fedex(batch):
    print("Loading Workbook...")
    workbook = openpyxl.load_workbook("FinalShippingTemplate.xlsx")
    workbook2 = openpyxl.load_workbook("AutoBatchPrintTemplate.xlsx")
    sheet = workbook["FedEx"]
    address = workbook2["Address"]
    row = 19
    arow = 2
    count = 1
    print('got', batch)
    print(f'{batch.num_packages} rows expected')

    process_addresses = 'no'
    for i in range(0, int(batch.num_packages)):

        batch.Billing = batch.Billing.lower() if batch.Billing is not None else batch.Billing
        batch.residential = batch.residential.lower() if batch.residential is not None else batch.residential
        batch.Description = batch.Description.lower() if batch.Description is not None else batch.Description
        batch.Method = batch.Method


        # -------------------------------------VERIFY ADDRESS -------------------------------------------------------
        address1 = address.cell(row=arow, column=2).value  # Address
        address2 = address.cell(row=arow, column=3).value if address.cell(row=arow, column=3).value is not None else ''
        city = address.cell(row=arow, column=5).value  # City
        country = country_search(address.cell(row=arow, column=8).value)  # search, convert the country to 2-letter code
        state = us_state_to_abbrev.get(address.cell(row=arow, column=6).value.strip().lower(),
                                       address.cell(row=arow,
                                                    column=6).value.strip()) if country == 'US' else (
            address.cell(row=arow, column=6).value.strip() if address.cell(row=arow,
                                                                           column=6).value is not None else '')
        zipcode = address.cell(row=arow, column=7).value if country != 'US' else fix_zip(
            address.cell(row=arow, column=7).value)

        # runs US address verification only if the address is in the US

        if process_addresses.lower() == 'yes':
            found = True
            if country == 'US':
                try:
                    vfstreet, vfstreet2, vfcity, vfstate, vfzipcode = verify_address_US(address1, address2, city, state,
                                                                                zipcode)
                    #  print(vfstreet)
                    if vfstreet is not None:
                        address1 = vfstreet
                        address2 = vfstreet2
                        city = vfcity
                        state = vfstate
                        zipcode = vfzipcode
                        print("Address Found")
                    else:
                        found = False
                except:
                    pass

            if not found:
                print("Address Unverified")

        if address.cell(row=arow, column=14).value is None:  # custom sort identifier
            sheet[f'A{row}'] = f'{batch.PackageId}-{str(count).zfill(4)}'
            sheet[f'B{row}'] = f'{batch.PackageId}-{str(count).zfill(4)}'
        else:
            sheet[f'A{row}'] = f'{batch.Ref1}'
            sheet[f'B{row}'] = f'({address.cell(row=arow, column=14).value}){batch.PackageId}-{str(count).zfill(4)}'

        sheet[f'C{row}'] = address.cell(row=arow, column=1).value  # ShipToName
        sheet[f'D{row}'] = ''
        sheet[f'E{row}'] = address1
        sheet[f'F{row}'] = address2
        sheet[f'G{row}'] = city
        country = country_search(address.cell(row=arow, column=8).value)  # Search, Convert country to 2-letter code
        sheet[f'I{row}'] = zipcode
        sheet[f'H{row}'] = state
        sheet[f'J{row}'] = country  # ship to country
        sheet[f'K{row}'] = batch.Weight
        sheet[f'L{row}'] = batch.Length
        sheet[f'M{row}'] = batch.Width
        sheet[f'N{row}'] = batch.Height
        sheet[f'O{row}'] = batch.SF_Company
        sheet[f'P{row}'] = 'Shipper'
        sheet[f'Q{row}'] = batch.SF_Address1
        sheet[f'R{row}'] = batch.SF_Address2
        sheet[f'S{row}'] = batch.SF_City
        sheet[f'T{row}'] = batch.SF_State
        sheet[f'U{row}'] = batch.SF_Zip

        sheet[f'V{row}'] = batch.AccountNum
        sheet[f'W{row}'] = '1' if batch.Billing.lower() == 'prepaid' else '3'
        sheet[f'X{row}'] = '1'
        sheet[f'Y{row}'] = '1'

        if search('fedex connect', str(batch.Method.lower())):
            sheet[f'Z{row}'] = 'FICP'
        if search('priority', str(batch.Method.lower())):
            sheet[f'Z{row}'] = '03'
        if search('international priority', str(batch.Method.lower())):
            sheet[f'Z{row}'] = '01'
        if search('overnight', str(batch.Method.lower())):
            sheet[f'Z{row}'] = '01'
        elif search('international economy', str(batch.Method.lower())):
            sheet[f'Z{row}'] = '03'
        elif search('expedited', str(batch.Method.lower())):
            sheet[f'Z{row}'] = '01'
        elif search('2 day', str(batch.Method.lower())):
            sheet[f'Z{row}'] = '03'
        elif search('home', str(batch.Method.lower())):
            sheet[f'Z{row}'] = '03'
        elif search('standard', str(batch.Method.lower())):
            sheet[f'Z{row}'] = '05'
        elif search('ground', str(batch.Method.lower())):
            sheet[f'Z{row}'] = '92'

        sheet[f'AA{row}'] = '0'
        sheet[f'AB{row}'] = '1' if batch.Billing == 'prepaid' else '3'
        sheet[f'AC{row}'] = '' if batch.Billing == 'prepaid' else batch.AccountNum
        sheet[f'AD{row}'] = phone_number_format(str(address.cell(row=arow, column=10).value)) if address.cell(row=arow, column=10).value is not None else '' # ship to phone
        sheet[f'AE{row}'] = batch.Description if country != 'US' else ''
        sheet[f'AF{row}'] = batch.Description if country != 'US' else ''
        sheet[f'AG{row}'] = batch.Origin
        sheet[f'AH{row}'] = '1'
        sheet[f'AI{row}'] = 'EA'
        sheet[f'AJ{row}'] = batch.Price if country != 'US' else ''
        sheet[f'AK{row}'] = batch.Price if country != 'US' else ''
        # sheet[f'AL{row}'] = '112669334'
        # sheet[f'AM{row}'] = 'Gift'
        # sheet[f'AN{row}'] = 'Not For Sale'
        sheet[f'AO{row}'] = batch.Harmonized_Code

        row += 1
        arow += 1
        count += 1

    workbook.save("FinalShippingTemplate.xlsx")
    print("Finished")

def fill_data_usps(batch):
    print("Loading Workbook...")
    workbook = openpyxl.load_workbook("FinalShippingTemplate.xlsx")
    workbook2 = openpyxl.load_workbook("AutoBatchPrintTemplate.xlsx")
    sheet = workbook["USPS"]
    address = workbook2["Address"]
    row = 4
    arow = 2
    count = 1
    print('got', batch)
    print(f'{batch.num_packages} rows expected')

    process_addresses = 'no'

    for i in range(0, int(batch.num_packages)):
        global vfstreet
        batch.Billing = batch.Billing.lower() if batch.Billing is not None else batch.Billing
        batch.residential = batch.residential.lower() if batch.residential is not None else batch.residential
        batch.Description = batch.Description.lower() if batch.Description is not None else batch.Description
        batch.Method = batch.Method.lower() if batch.Method is not None else batch.Method

        # -------------------------------------VERIFY ADDRESS -------------------------------------------------------
        address1 = address.cell(row=arow, column=2).value  # Address
        address2 = address.cell(row=arow, column=3).value if address.cell(row=arow, column=3).value is not None else ''
        city = address.cell(row=arow, column=5).value  # City
        country = country_search(address.cell(row=arow, column=8).value)  # search, convert the country to 2-letter code
        state = us_state_to_abbrev.get(address.cell(row=arow, column=6).value.strip().lower(),
                                       address.cell(row=arow,
                                                    column=6).value.strip()) if country == 'US' else (
            address.cell(row=arow, column=6).value.strip() if address.cell(row=arow,
                                                                           column=6).value is not None else '')
        zipcode = address.cell(row=arow, column=7).value if country != 'US' else fix_zip(
            address.cell(row=arow, column=7).value)

        # runs US address verification only if the address is in the US

        vfstreet, vfstreet2, vfcity, vfstate, vfzipcode = '', '', '', '', ''

        if process_addresses.lower() == 'yes':
            if country == 'US':
                found = False
                try:
                    vfstreet, vfstreet2, vfcity, vfstate, vfzipcode = verify_address_US(address1, address2, city, state,
                                                                                        zipcode)
                    if vfstreet and vfstreet2 and vfcity and vfstate and vfzipcode != '':
                        address1 = vfstreet
                        address2 = vfstreet2
                        city = vfcity
                        state = vfstate
                        zipcode = vfzipcode
                        print("Address Found")
                        found = True
                except:
                    print("EXCEPTION ----------------------------------")

                # color unverified rows
                if not found:
                    sheet[f'F{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'G{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'H{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'I{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'J{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'L{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'M{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'N{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'O{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))

                    sheet[f'P{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'Q{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'R{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'S{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'T{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'U{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))

                    sheet[f'V{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'W{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'X{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'Y{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))

                    sheet[f'Z{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AA{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AC{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AD{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'AE{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
        # ------------------------------------ WRITING THE DATA--------------------------------------------------------


        # print(f"vfstreet: {vfstreet}")
        sheet[f'F{row}'] = f'{batch.PackageId}-{str(count).zfill(4)}'  # Shipment Key
        sheet[f'G{row}'] = address.cell(row=arow, column=1).value  # ShipToName
        sheet[f'H{row}'] = ''
        sheet[f'I{row}'] = address1
        sheet[f'J{row}'] = address2
        sheet[f'L{row}'] = city
        country = country_search(address.cell(row=arow, column=8).value)  # search, convert the country to 2-letter code
        sheet[f'M{row}'] = state
        sheet[f'N{row}'] = zipcode
        sheet[f'O{row}'] = country

        sheet[f'P{row}'] = batch.SF_Company
        sheet[f'Q{row}'] = batch.SF_Address1
        sheet[f'R{row}'] = batch.SF_Address2
        sheet[f'S{row}'] = batch.SF_City
        sheet[f'T{row}'] = batch.SF_State
        sheet[f'U{row}'] = batch.SF_Zip

        sheet[f'V{row}'] = batch.Length
        sheet[f'W{row}'] = batch.Width
        sheet[f'X{row}'] = batch.Height
        sheet[f'Y{row}'] = batch.Weight

        sheet[f'Z{row}'] = batch.Description if country != 'US' else ''
        sheet[f'AA{row}'] = batch.Method
        sheet[f'AC{row}'] = batch.Description if country != 'US' else ''
        sheet[f'AD{row}'] = batch.Price if country != 'US' else ''
        sheet[f'AE{row}'] = batch.Price if country != 'US' else ''

        row += 1
        arow += 1
        count += 1

    workbook.save("FinalShippingTemplate.xlsx")
    print("Finished")

def fill_data_prefship(batch):
    print("Loading Workbook...")
    workbook = openpyxl.load_workbook("FinalShippingTemplate.xlsx")
    workbook2 = openpyxl.load_workbook("AutoBatchPrintTemplate.xlsx")
    sheet = workbook["PrefShip"]
    address = workbook2["Address"]
    row = 2
    arow = 2
    count = 1
    print('got', batch)
    print(f'{batch.num_packages} rows expected')

    process_addresses = 'no'

    for i in range(0, int(batch.num_packages)):
        batch.Description = batch.Description.lower() if batch.Description is not None else batch.Description
        batch.Method = batch.Method.lower() if batch.Method is not None else batch.Method

        address1 = address.cell(row=arow, column=2).value  # Address
        address2 = address.cell(row=arow, column=3).value if address.cell(row=arow, column=3).value is not None else ''
        city = address.cell(row=arow, column=5).value  # City
        country = country_search(address.cell(row=arow, column=8).value)  # search, convert the country to 2-letter code
        state = us_state_to_abbrev.get(address.cell(row=arow, column=6).value.strip().lower(),
                                       address.cell(row=arow,
                                                    column=6).value.strip()) if country == 'US' else (
            address.cell(row=arow, column=6).value.strip() if address.cell(row=arow,
                                                                           column=6).value is not None else '')
        zipcode = address.cell(row=arow, column=7).value if country != 'US' else fix_zip(
            address.cell(row=arow, column=7).value)

        if process_addresses.lower() == 'yes':
            if country == 'US':
                found = False
                try:
                    vfstreet, vfstreet2, vfcity, vfstate, vfzipcode = verify_address_US(address1, address2, city, state,
                                                                                        zipcode)
                    if vfstreet and vfstreet2 and vfcity and vfstate and vfzipcode != '':
                        address1 = vfstreet
                        address2 = vfstreet2
                        city = vfcity
                        state = vfstate
                        zipcode = vfzipcode
                        print("Address Found")
                        found = True
                except:
                    print("EXCEPTION ----------------------------------")

                # color unverified rows
                if not found:
                    sheet[f'F{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'G{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'H{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'I{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'J{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'L{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'M{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'N{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'O{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))

                    sheet[f'P{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'Q{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'R{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'S{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'T{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'U{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))

                    sheet[f'V{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'W{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'X{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))
                    sheet[f'Y{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))

                    sheet[f'Z{row}'].fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='FF0000'))

        if address.cell(row=arow, column=14).value is None:  # custom sort identifier
            sheet[f'A{row}'] = f'{batch.PackageId}-{str(count).zfill(4)}'
            sheet[f'B{row}'] = f'{batch.PackageId}-{str(count).zfill(4)}'
        else:
            sheet[f'A{row}'] = f'({address.cell(row=arow, column=14).value}){batch.PackageId}-{str(count).zfill(4)}'
            sheet[f'B{row}'] = f'({address.cell(row=arow, column=14).value}){batch.PackageId}-{str(count).zfill(4)}'

        sheet[f'C{row}'] = address.cell(row=arow, column=1).value.strip() if address.cell(row=arow,
                                                                                          column=1).value is not None else ''  # name
        sheet[f'D{row}'] = str(address.cell(row=arow, column=2).value).strip() if address.cell(row=arow,
                                                                                               column=2).value is not None else ''  # Address
        sheet[f'E{row}'] = address.cell(row=arow, column=3).value.strip() if address.cell(row=arow,
                                                                                          column=3).value is not None else ''  # Address2
        sheet[f'G{row}'] = address.cell(row=arow, column=5).value.strip() if address.cell(row=arow,
                                                                                          column=5).value is not None else ''  # City
        country = country_search(address.cell(row=arow, column=8).value)  # search, convert the country to 2-letter code
        sheet[f'H{row}'] = us_state_to_abbrev.get(address.cell(row=arow, column=6).value.strip().lower(),
                                                  address.cell(row=arow,
                                                               column=6).value.strip()) if country == 'US' else (
            address.cell(row=arow, column=6).value.strip() if address.cell(row=arow,
                                                                           column=6).value is not None else '')  # state
        sheet[f'I{row}'] = address.cell(row=arow, column=7).value if country != 'US' else fix_zip(
            address.cell(row=arow, column=7).value)
        sheet[f'J{row}'] = country
        sheet[f'K{row}'] = phone_number_format(str(address.cell(row=arow, column=10).value)) if address.cell(row=arow, column=10).value is not None else '' # ship to phone

        sheet[f'M{row}'] = batch.SF_Company
        sheet[f'N{row}'] = batch.SF_Address1
        sheet[f'O{row}'] = batch.SF_Address2
        sheet[f'P{row}'] = batch.SF_City
        sheet[f'Q{row}'] = batch.SF_State
        sheet[f'R{row}'] = batch.SF_Zip
        sheet[f'S{row}'] = 'US'

        sheet[f'T{row}'] = batch.Weight
        sheet[f'U{row}'] = batch.Length
        sheet[f'V{row}'] = batch.Width
        sheet[f'W{row}'] = batch.Height
        sheet[f'X{row}'] = batch.Price
        sheet[f'Y{row}'] = '1'
        sheet[f'Z{row}'] = batch.Description

        row += 1
        arow += 1
        count += 1

    workbook.save("FinalShippingTemplate.xlsx")
    print("Finished")

def saveTemplate(wbook):
    name = input('Enter Template name: ')
    wbook.save(rf'\\fpisql\Apps\BatchPrints\Bamco\{name}.xlsx')


def main():
    bp = Batch("AutoBatchPrintTemplate.xlsx")
    bp.display()
    if search('ups', str(bp.Carrier).lower()):
        fill_data_ups(bp)
    elif search('fedex', str(bp.Carrier).lower()):
        fill_data_fedex(bp)
    elif search('usps', str(bp.Carrier).lower()):
        fill_data_usps(bp)
    elif search('prefship', str(bp.Carrier).lower()):
        fill_data_prefship(bp)
    else:
        print("No shipping method specified...")

    os.system("start EXCEL.EXE FinalShippingTemplate.xlsx")


# ------------------------ RUN --------------------------------

if __name__ == "__main__":
    main()
